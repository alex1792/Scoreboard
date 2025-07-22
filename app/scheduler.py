import pandas as pd
import random
from collections import Counter
from itertools import combinations
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# read files
def read_file(fname='24FALL_USC_OPEN_TEST.xlsx'):
    f = pd.read_excel(fname, engine='openpyxl')
    return f

# split player names, compute weights
# match = [team1_names, team2_names, category, flight, weight]
def get_match_info(f):
    """
    This function is used to get the match info from the file
    It will return a list of matches with the following format:
    [player1s, player2s, category, flight, weight]

    player1s: a list of player names in team 1
    player2s: a list of player names in team 2
    category: the category of the match (MS, WS, MD, WD, XD)
    flight: the group of the match (A, B, C...)
    weight: the weight of the match, computed by the number of the remaining matches of each player
    """
    # get all player's name
    player1s = f['Team1'].tolist()
    player2s = f['Team2'].tolist()
    categories = f['Category'].tolist()
    flights = f['Flight'].tolist()
            
    # analyze each player's name and their total games
    players = {}
    matches = []
    for player1, player2, category, flight in zip(player1s, player2s, categories, flights):
        # men's single, women's single
        if category == "MS" or category == "WS":
            players[player1] = players.get(player1, 0) + 1
            players[player2] = players.get(player2, 0) + 1
            matches.append([[player1], [player2], category, flight])
        # men's doubles, women's double, mixed doubles
        else:
            p1s = [p.strip() for p in player1.strip().split('/')]
            # print(p1s)
            for p in p1s:
                players[p] = players.get(p, 0) + 1
            p2s = [p.strip() for p in player2.strip().split('/')]
            # print(p2s)
            for p in p2s:
                players[p] = players.get(p, 0) + 1
            matches.append([p1s, p2s, category, flight])
            
    # compute weight of each match
    for m in matches:
        if m[2] == "MS" or m[2] == "WS":
            m.append(players[m[0][0]] + players[m[1][0]])
        else:
            total = 0
            for p in m[0]:
                total += players[p]
            for p in m[1]:
                total += players[p]
            m.append(total / 2)
    return matches

# Calculate the weight of each match
def count_player_matches(matches):
    counter = Counter()
    for m in matches:
        for p in m[0] + m[1]:
            counter[p] += 1
    return counter

def calc_weight(match, player_counts):
    p1s, p2s = match[0], match[1]
    if len(p1s) == 1 and len(p2s) == 1:  # single
        return player_counts[p1s[0]] + player_counts[p2s[0]]
    else:  # doubles
        return (sum(player_counts[p] for p in p1s) + sum(player_counts[p] for p in p2s)) / 2

def update_weights(matches, player_counts):
    for match in matches:
        match[4] = calc_weight(match, player_counts)
    return matches

def update_remaining_weights(matches):
    """update remaining weights based on each player's remaining matches"""
    player_counts = {}
    # calculate the remaining match for each player
    for match in matches:
        for p in set(match[0]) | set(match[1]):
            player_counts[p] = player_counts.get(p, 0) + 1
    
    # update weight of each match (match weight = weight of team 1 + weight of team 2)
    updated_matches = []
    for match in matches:
        players = set(match[0]) | set(match[1])
        new_weight = sum(player_counts.get(p, 0) for p in players)
        updated_match = (*match[:-1], new_weight)
        updated_matches.append(updated_match)
    return updated_matches

# print sorted match
def print_sorted_match(matches):
    sorted_match = sorted(matches, key=lambda x:x[4], reverse=True)
    print('='*100)
    print('Sorted match:')
    for m in sorted_match:
        print(m)
    print('='*100)

# GREEDY ALGORITHM
def scheduler(matches, total_court, monte_carlo=False):
    batches = []
    prev_batch_players = set()
    
    while matches:
        # weight
        matches_sorted = sorted(matches, key=lambda m: -m[4])  
        
        batch = []
        players_in_batch = set()
        remaining = []
        candidates = []  
        
        # collect the valid matches, and put them into the candidate pool 
        for match in matches_sorted:
            p1s, p2s = set(match[0]), set(match[1])
            if (not (p1s & players_in_batch) and \
                not (p2s & players_in_batch) and \
                not (p1s & prev_batch_players) and \
                not (p2s & prev_batch_players)):
                candidates.append(match)
            else:
                remaining.append(match)
        
        # Monte Carlo, randomly fill the batch
        while len(batch) < total_court and candidates:
            if monte_carlo:
                # randomly choose one match
                selected = random.choice(candidates)
            else:
                # original greedy algo, pick the highest weight match
                selected = candidates[0]  
            
            batch.append(selected)
            candidates.remove(selected)
            # update players in batch
            players_in_batch.update(selected[0])
            players_in_batch.update(selected[1])
            
            # dynamically update the candidate pool: remove confilct match, add new candidates
            new_candidates = []
            for candidate in candidates:
                cp1, cp2 = set(candidate[0]), set(candidate[1])
                if not (cp1 & players_in_batch or cp2 & players_in_batch):
                    new_candidates.append(candidate)
                else:
                    remaining.append(candidate)
            candidates = new_candidates
        
        # if this batch can't schedule any match, then return
        if not batch:
            break
            
        batches.append(batch)
        prev_batch_players = players_in_batch
        matches = remaining + candidates  # the remaining match will go into next round
        matches = update_remaining_weights(matches)  # update weights

    
    
    return fit_remaining_into_batch(batches, matches, total_court)

# fit the remaining match into batch that are not full
def fit_remaining_into_batch(batches, remaining, total_court, monte_carlo=True):
    if not remaining:
        return batches, remaining, 0
    
    player_last_batch = {}
    total_inserted = 0

    # initialize the last batch of each player
    for batch_idx, batch in enumerate(batches):
        for match in batch:
            for player in set(match[0]) | set(match[1]):
                player_last_batch[player] = batch_idx

    while True:
        inserted_this_round = False

        # update weights and sort by weight (large -> small)
        remaining = update_remaining_weights(remaining)
        remaining = sorted(remaining, key=lambda x: -x[4])

        for batch_idx, batch in enumerate(batches):
            # get the current batch players name set
            current_batch_players = set()
            for m in batch:
                current_batch_players.update(m[0])
                current_batch_players.update(m[1])
            
            # collect all insertable match, and save them into the candidate pool
            candidate_matches = []
            for j, match in enumerate(remaining):
                players = set(match[0]) | set(match[1])

                # check if players of potential inserting match are in current batch or not
                if players & current_batch_players:
                    continue
                else:
                    candidate_matches.append((j, match))
            
            # if the candidate is not empty, then randomly insert
            if candidate_matches:
                if monte_carlo:
                    selected_index, selected_match = random.choice(candidate_matches)
                else:
                    # default choose the hightest weight
                    selected_index, selected_match = candidate_matches[0]
                
                batch.append(selected_match)
                
                players = set(selected_match[0]) | set(selected_match[1])
                current_batch_players.update(players)
                for p in players:
                    player_last_batch[p] = batch_idx
                remaining.pop(selected_index)
                total_inserted += 1
                inserted_this_round = True
                break  # break after insertion

        if not inserted_this_round:
            break  # if there's no insertion, then break

    # handle the remaining match, concat at the end
    if remaining:
        new_batches, remaining = schedule_remaining_simple(remaining, total_court)
        batches.extend(new_batches)
        total_inserted += sum(len(b) for b in new_batches)

    # print(f'insert {total_inserted} matches into schedule')
    return batches, remaining, total_inserted

# schedule the remaining match into batch that are not full
def schedule_remaining_simple(remaining, total_court):
    batches = []
    while remaining:
        batch = []
        players_in_batch = set()
        i = 0
        while i < len(remaining) and len(batch) < total_court:
            match = remaining[i]
            players = set(match[0]) | set(match[1])
            if not (players & players_in_batch):  # only check current_batch
                batch.append(match)
                players_in_batch.update(players)
                remaining.pop(i)
            else:
                i += 1
        batches.append(batch)
    return batches, remaining

# annotate consecutive players
def annotate_consecutive_players(batches):
    prev_batch_players = set()
    new_batches = []
    for batch in batches:
        new_batch = []
        current_batch_players = set()
        for match in batch:
            players = set(match[0]) | set(match[1])
            # find consecutive players (if current player plays in this and previous batch)
            consecutive_players = [p for p in players if p in prev_batch_players]
            marked_match = list(match) + [consecutive_players]
            new_batch.append(marked_match)
            current_batch_players.update(players)
        prev_batch_players = current_batch_players  # update previous batch players
        new_batches.append(new_batch)
    return new_batches

# backtracking method
def can_batch(batch, prev_batch_players):
    players_in_batch = set()
    for match in batch:
        p1s, p2s = set(match[0]), set(match[1])
        if (p1s | p2s) & players_in_batch:
            return False
        if (p1s | p2s) & prev_batch_players:
            return False
        players_in_batch.update(p1s)
        players_in_batch.update(p2s)
    return True

# backtracking method
def backtracking_scheduler(matches, total_court, batches=None, prev_batch_players=None):
    # print('called')
    if batches is None:
        batches = []
    if prev_batch_players is None:
        prev_batch_players = set()

    if not matches:
        return batches  # all matches are scheduled

    if len(matches) >= total_court:
        for batch in combinations(matches, total_court):
            if can_batch(batch, prev_batch_players):
                remaining = [m for m in matches if m not in batch]
                players_in_batch = set()
                for match in batch:
                    players_in_batch.update(match[0])
                    players_in_batch.update(match[1])
                result = backtracking_scheduler(remaining, total_court, batches + [list(batch)], players_in_batch)
                if result is not None:
                    return result
    else:
        # last batch, size may not to be fixed
        for batch_size in range(len(matches), 0, -1):
            for batch in combinations(matches, batch_size):
                if can_batch(batch, prev_batch_players):
                    remaining = [m for m in matches if m not in batch]
                    players_in_batch = set()
                    for match in batch:
                        players_in_batch.update(match[0])
                        players_in_batch.update(match[1])
                    result = backtracking_scheduler(remaining, total_court, batches + [list(batch)], players_in_batch)
                    if result is not None:
                        return result

    return None  # match can't be scheduled

# write schedule to excel file
def write_schedule(batches, total_court, filename):
    rows = []
    all_consecutive_players = []  # 收集所有連續出場的選手
    
    for batch_idx, batch in enumerate(batches, 1):
        # 為每個 batch 創建固定數量的 court
        batch_rows = []
        
        # 添加實際的比賽
        for court_idx, match in enumerate(batch, 1):
            consecutive_players = match[5] if len(match) > 5 else []
            consecutive_str = ", ".join(consecutive_players) if consecutive_players else ""
            
            # 收集所有連續出場的選手
            all_consecutive_players.extend(consecutive_players)
            
            # 獲取更多比賽信息
            category = match[2]
            flight = match[3]
            player1s = ' / '.join(match[0])
            player2s = ' / '.join(match[1])
            
            # 判斷比賽類型
            if len(match[0]) == 1 and len(match[1]) == 1:
                match_type = "Single"
            else:
                match_type = "Double"
            
            batch_rows.append({
                'Round': batch_idx,
                'Court': f"Court {court_idx}",
                'Match_Type': match_type,
                'Category': category,
                'Group': flight,
                'Player1/Team1': player1s,
                'Player2/Team2': player2s,
                'Consecutive_Players': consecutive_str,
                'Status': 'Scheduled',
                'Score1': 0,
                'Score2': 0,
                'Umpire': '',
                'Notes': ''
            })
        
        # 填充空行以達到 total_court 數量
        for court_idx in range(len(batch) + 1, total_court + 1):
            batch_rows.append({
                'Round': batch_idx,
                'Court': f"Court {court_idx}",
                'Match_Type': '',
                'Category': '',
                'Group': '',
                'Player1/Team1': '',
                'Player2/Team2': '',
                'Consecutive_Players': '',
                'Status': '',
                'Score1': '',
                'Score2': '',
                'Umpire': '',
                'Notes': ''
            })
        
        # 將這個 batch 的所有行添加到總行列表
        rows.extend(batch_rows)

    # 計算受影響選手統計
    from collections import Counter
    player_counts = Counter(all_consecutive_players)
    total_affected_players = len(player_counts)
    
    # 添加統計信息行
    rows.append({})  # 空行
    rows.append({
        'Round': 'Stats',
        'Court': '',
        'Match_Type': '',
        'Category': '',
        'Group': '',
        'Player1/Team1': '',
        'Player2/Team2': '',
        'Consecutive_Players': '',
        'Status': '',
        'Score1': '',
        'Score2': '',
        'Umpire': '',
        'Notes': ''
    })
    
    # 添加受影響人數總計
    rows.append({
        'Round': 'Total Affected Players',
        'Court': '',
        'Match_Type': '',
        'Category': '',
        'Group': '',
        'Player1/Team1': '',
        'Player2/Team2': '',
        'Consecutive_Players': '',
        'Status': '',
        'Score1': '',
        'Score2': '',
        'Umpire': '',
        'Notes': total_affected_players
    })
    
    # 添加每個受影響選手的次數
    for player, count in player_counts.most_common():  # 按次數降序排列
        rows.append({
            'Round': '',
            'Court': '',
            'Match_Type': '',
            'Category': '',
            'Group': '',
            'Player1/Team1': player,
            'Player2/Team2': '',
            'Consecutive_Players': '',
            'Status': '',
            'Score1': '',
            'Score2': '',
            'Umpire': '',
            'Notes': f"Consecutive {count} times"
        })

    df = pd.DataFrame(rows)
    df.to_excel(filename, index=False, sheet_name='MatchSchedule')

    # 使用 openpyxl 添加顏色標記
    wb = load_workbook(filename)
    ws = wb['MatchSchedule']
    
    # 黃色標記連續出場的選手
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    # 綠色標記有比賽的行
    green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    
    # 藍色標記統計信息行
    blue_fill = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")
    
    # 紅色標記受影響選手統計
    red_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column), 2):
        consecutive_cell = row[7]  # Consecutive_Players 欄位
        category_cell = row[3]     # Category 欄位
        round_cell = row[0]        # Round 欄位
        
        # 檢查是否為統計信息行
        if round_cell.value == 'Stats':
            for cell in row:
                cell.fill = blue_fill
        elif round_cell.value == 'Total Affected Players':
            for cell in row:
                cell.fill = red_fill
        elif round_cell.value == '' and row[5].value and 'Consecutive' in str(row[11].value or ''):
            # 受影響選手詳細統計行
            for cell in row:
                cell.fill = red_fill
        elif consecutive_cell.value:  # 有連續出場選手
            for cell in row:
                cell.fill = yellow_fill
        elif category_cell.value:   # 有比賽但無連續出場
            for cell in row:
                cell.fill = green_fill

    wb.save(filename)

# print final schedule
def print_final_schedule(batches, remaining):
    print('='*100)
    for batch_idx, batch in enumerate(batches, 1):
        for match in batch:
            print(f'{batch_idx}\t{match}')
    print('='*100)
    print('='*100)
    print('these match can not be assigned:')
    for r in remaining:
        print(r)
    print('='*100)

# generate schedule using greedy algorithm
def generate_schedule(f, total_court, output_fname):
    # f = read_file(fname)
    
    matches = get_match_info(f)
    
    # print_sorted_match(matches)
    
    batches, remaining, total_insert = scheduler(matches, total_court)

    batches = annotate_consecutive_players(batches)
    
    write_schedule(batches, total_court, output_fname)
    
    # print_final_schedule(batches, remaining)

# generating schedule using backtracking(TLE)
def generate_schedule_backtracking(fname, total_court):
    f = read_file(fname)
    matches = get_match_info(f)

    schedule = backtracking_scheduler(matches, total_court)
    write_schedule(schedule, total_court, 'backtracking_schedule.xlsx')
    if schedule:
        for i, batch in enumerate(schedule):
            print(f"Round {i+1}: {batch}")
    else:
        print("No feasible schedule found.")

def get_match_info_from_database(matches):
    """
    Get matches from database and transform to the format of scheduler.
    
    Args:
        matches: match filtered by tournament_id
    
    Returns:
        list: scheduler format [[player1_list], [player2_list], category, group, weight]
    """
    from .models import Group  # 添加這個導入
    
    # 分析每個選手的比賽數量
    players = {}
    converted_matches = []
    
    for match in matches:
        # 獲取 group 信息
        group = Group.query.get(match.group_id) if match.group_id else None
        group_name = group.name if group else 'A'
        
        # 獲取選手姓名
        if match.event_type in ['MS', 'WS']:  # 單打
            player1_name = match.player1_name or (match.player1.get_full_name() if match.player1 else "Unknown")
            player2_name = match.player2_name or (match.player2.get_full_name() if match.player2 else "Unknown")
            
            # 更新選手比賽數量
            players[player1_name] = players.get(player1_name, 0) + 1
            players[player2_name] = players.get(player2_name, 0) + 1
            
            converted_matches.append([
                [player1_name], 
                [player2_name], 
                match.event_type, 
                group_name,  # 使用正確獲取的 group_name
                1  # 預設權重
            ])
        else:  # 雙打
            # 獲取選手姓名
            team1_player1 = match.team1_player1_name or (match.team1_player1.get_full_name() if match.team1_player1 else "Unknown")
            team1_player2 = match.team1_player2_name or (match.team1_player2.get_full_name() if match.team1_player2 else "Unknown")
            team2_player1 = match.team2_player1_name or (match.team2_player1.get_full_name() if match.team2_player1 else "Unknown")
            team2_player2 = match.team2_player2_name or (match.team2_player2.get_full_name() if match.team2_player2 else "Unknown")
            
            # 更新選手比賽數量
            for player in [team1_player1, team1_player2, team2_player1, team2_player2]:
                players[player] = players.get(player, 0) + 1
            
            converted_matches.append([
                [team1_player1, team1_player2], 
                [team2_player1, team2_player2], 
                match.event_type, 
                group_name,  # 使用正確獲取的 group_name
                1  # 預設權重
            ])
    
    # 計算每個比賽的權重
    for match in converted_matches:
        if len(match[0]) == 1 and len(match[1]) == 1:  # 單打
            weight = players[match[0][0]] + players[match[1][0]]
        else:  # 雙打
            total_weight = 0
            for player in match[0] + match[1]:
                total_weight += players[player]
            weight = total_weight / 2
        match[4] = weight  # 更新權重
    
    return converted_matches

def generate_schedule_from_database(matches, total_court, output_fname):
    """
    From database to generate schedule. No need to upload file.
    
    Args:
        matches: matches from database
        total_court: available courts in the arena
        output_fname: output file name
    """
    # 從資料庫對象獲取比賽信息
    match_data = get_match_info_from_database(matches)
    
    # 生成賽程表
    batches, remaining, total_insert = scheduler(match_data, total_court)
    
    # 標註連續出場的選手
    batches = annotate_consecutive_players(batches)
    
    # 寫入 Excel 檔案（保持與原本相同的格式）
    write_schedule(batches, total_court, output_fname)
    
    # 打印結果
    # print_final_schedule(batches, remaining)
    
    return {
        'batches': batches,
        'remaining': remaining,
        'total_inserted': total_insert
    }

# 更新現有的函數以使用新的基於資料庫的函數
def generate_schedule_for_tournament_from_matches(matches, total_court, output_fname):
    """
    From database to generate schedule. No need to upload file.
    
    Args:
        matches: matches from database
        total_court: available courts in the arena
        output_fname: output file name
    """
    return generate_schedule_from_database(matches, total_court, output_fname)  