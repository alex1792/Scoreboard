from ..models import Match, Group, db, Tournament, Schedule, ScheduleItem
from datetime import datetime, timedelta
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from collections import Counter
from .tournament_service import TournamentService

class TournamentScheduler:
    """use to generate schedule for a tournament"""
    def __init__(self, total_court):
        self.total_court = total_court
        self.scheduled_matches = [] # store previous batch of match
        self.completed_matches = set()
        self.all_matches = []  # store all matches

    def schedule_tournament(self, matches, tournament_id):
        """schedule the tournament, handle the incomplete batches and dependencies

        ***** the core schedule concept is to avoid players have more than 1 games at same scheduling batch *****
        ***** Under this coure concept, there might be matches that couldn't fit in the scheduleing batch   *****
        ***** For example, A player signed up for both single and double event. You can not assign both     *****
        ***** player A's single and double games at the same batch. Either pick singe or double match.       *****
        
        the pseudo algorithm is:
            1. group the match by round
            2. process the matches by round (ensure that round_1 matches are complete scheduled before processing the next round)
                a. initialize the candidate matches for each round (eg: round_1_all_candidates, round_2_all_candidates, ...)
                b. add matches in that round to the candidate matches in 2-a
                c. process the round_n_all_candidates
            3. handle with remaining matches
        
        """
        # ensure matches is a list
        if not isinstance(matches, (list, tuple)):
            matches = [matches]
        
        self.all_matches = matches
        self.scheduled_matches = []  # reset the scheduled matches
        
        # 1. group the matches by round
        matches_by_round = self._group_by_round(matches)
        
        # 2. process the matches by round - ensure Round 1 is completely scheduled before processing Round 2
        for round_num in sorted(matches_by_round.keys()):
            round_matches = matches_by_round[round_num]
            
            # 3. initialize the candidate matches for the round
            candidate_matches = set()
            
            # 4. add the matches in the round to the candidate matches
            for match in round_matches:
                if self._can_schedule_match(match):
                    candidate_matches.add(match)
            
            # 5. schedule all matches in the round
            while candidate_matches:
                batch_matches = self._schedule_batch(candidate_matches)
                if not batch_matches:
                    break
                
                # 6. update the candidate matches
                for match in batch_matches:
                    candidate_matches.discard(match)
                    if match not in self.scheduled_matches:
                        self.scheduled_matches.append(match)
                
                # 7. re-check the matches that can be scheduled in the round
                self._update_candidate_matches_for_round(candidate_matches, round_matches)
        
        # 8. final stage: fill the remaining matches
        self._fill_remaining_matches()

        if tournament_id:
            # Get the exact time of tournament
            tournament = Tournament.query.get(tournament_id)
            if tournament:
                start_time = tournament.start_time if tournament.start_time else '09:00'
                end_time = tournament.end_time if tournament.end_time else '18:00'
                match_duration = tournament.match_duration if tournament.match_duration else 30
                
                schedule_data = {
                    'start_time': start_time,
                    'end_time': end_time, 
                    'match_duration': match_duration,
                }
            else:
                # Use default time when the start_time is not found
                schedule_data = {
                    'start_time': '09:00',
                    'end_time': '18:00',
                    'match_duration': 30
                }
            # print(f"schedule_data: {schedule_data}")
            self.create_schedule(tournament_id, schedule_data)

    def _group_by_round(self, matches):
        """
        group the matches by round
        
        for example:
            elimination match - round 1, 2, 3, ... till final round
            round robin - only first round
        
        this function is to group matches by round
        matches_by_round = {
            "1", [r1_match_1, r1_match_2, ..., r1_match_n],
            "2", [r2_match1, r2_match_2, ..., r2_match_m],
            ...
            }
        """
        matches_by_round = {}
        
        for match in matches:
            round_num = match.round or 1  # if no round, set to 1
            if round_num not in matches_by_round:
                matches_by_round[round_num] = []
            matches_by_round[round_num].append(match)
        
        return matches_by_round
    
    def _get_match_by_id(self, match_id):
        """get the match by match_id"""
        if not match_id:
            return None
        
        for match in self.all_matches:
            if match.id == match_id:
                return match
        return None
    
    def _schedule_batch(self, candidate_matches):
        """schedule a batch of matches
        
        using the candidate_matches to pick the matches that will be scheduled in the schedule batch.

        For each match, calculate the weight based on the remaining match of both players using _calculate_weight()

        For instance: if match_1 is Player A vs Player B (single match) or Player A / Player B vs Player C / Player D
            1. find the remaining match of both Player A and B (Player C and D if match.category is doubles)
            2. Consider the resting time of each player, the longer resting time, the higher priority it should be scheduled
        
        """
        selected_matches = []
        selected_players = set()
        
        # calculate the weight and sort
        weighted_matches = []
        for match in candidate_matches:
            weight = self._calculate_weight(match)
            weighted_matches.append((match, weight))
        
        weighted_matches.sort(key=lambda x: x[1], reverse=True)
        
        # greedy selection
        for match, weight in weighted_matches:
            if len(selected_matches) >= self.total_court:
                break
            
            match_players = self._get_match_players(match)
            if not (match_players & selected_players):
                selected_matches.append(match)
                selected_players.update(match_players)
        
        return selected_matches

    def _calculate_weight(self, match):
        """
        calculate the weight of the match
        
        consider the following factors:
            1. remaining games (if the player has more games to play, then the match should be scheduled first)
            2. resting time (if the player has been resting a lot, then he should have a higher priority to be scheduled)
        """
        weight = 0
        
        # 1. remaining games (if the player has more games to play, then the match should be scheduled first)
        weight += self._compute_weight_for_remaining_games(match)
        
        # 2. resting time (if the player has been resting a lot, then he should have a higher priority to be scheduled)
        weight += self._compute_weight_for_resting_time(match)
        
        return weight
    
    def _compute_weight_for_remaining_games(self, match):
        """calculate the weight based on the remaining games"""
        weight = 0
        
        # get all players
        players = self._get_match_players(match)
        
        # calculate the remaining games for each player
        for player in players:
            remaining_games = self._get_remaining_games_for_player(player)
            weight += remaining_games * 10
        
        return weight
    
    def _get_match_players(self, match):
        """
        get all players in the match
        
        case 1: for single players, return two players
        case 2: for double players, return fourn players

        return a list of string 
        """
        players = set()
        
        if match.event_type in ['MD', 'WD', 'XD']:
            if match.team1_player1_name:
                players.add(match.team1_player1_name)
            if match.team1_player2_name:
                players.add(match.team1_player2_name)
            if match.team2_player1_name:
                players.add(match.team2_player1_name)
            if match.team2_player2_name:
                players.add(match.team2_player2_name)
        else:
            if match.player1_name:
                players.add(match.player1_name)
            if match.player2_name:
                players.add(match.player2_name)
        
        return players
    
    def _get_remaining_games_for_player(self, player_name):
        """
        get the remaining games for the player
        
        for example: player A has 5 more games not been scheduled, the function will return 5
        
        """
        remaining = 0
        
        for match in self.all_matches:
            if match.status != 'ended':
                players = self._get_match_players(match)
                if player_name in players:
                    remaining += 1
        
        return remaining
    
    def _compute_weight_for_resting_time(self, match):
        """
        calculate the weight based on the resting time (based on batch)
        
        this function will check all players in parameter 'match'
        from previous batch, if player has played last batch, then the match should have lower priority to be scheduled
        (make sure players have enough rest)
        """
        penalty = 0
        
        # get the players of the current match
        current_players = self._get_match_players(match)
        
        # check if the players have appeared in the previous batches
        for player in current_players:
            if player in self._get_players_from_previous_batches():
                penalty -= 100  # heavily penalize consecutive players
        
        return penalty

    def _get_players_from_previous_batches(self):
        """get all players from the previous batches"""
        players = set()
        
        # get players from the scheduled matches
        for match in self.scheduled_matches:
            match_players = self._get_match_players(match)
            players.update(match_players)
        
        return players

    def _update_candidate_matches_for_round(self, candidate_matches, round_matches):
        """
        update the candidate matches for the round
        

        After putting some matches into curret_batch, need to update the remaining candidates
        ensure that candidate_matches only contain unscheduled matches
        """
        for match in round_matches:
            if match not in self.scheduled_matches and self._can_schedule_match(match):
                candidate_matches.add(match)

    def _fill_remaining_matches(self):
        """
        fill the remaining matches, minimize the number of affected players
        
        since there are some cases that we can not fit all match into the schedule table
        that satisfy the condition that players won't play back to back.

        Under the circumstances, we have to minimize the number of affected players.
        For each batch, there might be empty slots, the algo will schedule match into these slots.
        Once the all slots are scheduled, we have to create new batch to fit the unscheduled matches.
        """
        # get all the unscheduled matches
        scheduled_match_ids = {match.id for match in self.scheduled_matches}
        remaining_matches = [match for match in self.all_matches if match.id not in scheduled_match_ids]
        
        if not remaining_matches:
            return
        
        # print(f"Found {len(remaining_matches)} remaining matches to schedule")
        
        # sort the remaining matches by round
        remaining_matches.sort(key=lambda x: (x.round or 1, x.match_number or 1))
        
        # try to fill the incomplete batches
        self._fill_incomplete_batches(remaining_matches)
        
        # if there are still remaining matches, create new batches
        remaining_match_ids = {match.id for match in remaining_matches}
        still_remaining = [match for match in self.all_matches if match.id in remaining_match_ids]
        
        if still_remaining:
            # print(f"Creating new batches for {len(still_remaining)} remaining matches")
            self._create_new_batches_for_remaining(still_remaining)
        
        # final check
        final_remaining = [match for match in self.all_matches if match not in self.scheduled_matches]
        if final_remaining:
            # print(f"Warning: {len(final_remaining)} matches still not scheduled")
            for match in final_remaining:
                print(f"  - Match {match.id}: {match.player1_name} vs {match.player2_name} (Round {match.round})")

    def _fill_incomplete_batches(self, remaining_matches):
        """fill the incomplete batches"""
        # reorganize the scheduled matches into batches
        batches = self._organize_matches_into_batches()
        
        for batch_num, batch in batches.items():
            if len(batch) < self.total_court:
                # this batch is not full, try to fill the matches
                self._fill_batch_with_remaining(batch_num, batch, remaining_matches)

    def _fill_batch_with_remaining(self, batch_idx, batch, remaining_matches):
        """fill the remaining matches in the specified batch"""
        batch_players = set()
        for match_info in batch:
            match = match_info['match']
            match_players = self._get_match_players(match)
            batch_players.update(match_players)
        
        # find the matches that can be filled
        fillable_matches = []
        for match in remaining_matches:
            if self._can_fill_match_in_batch(match, batch_players):
                fillable_matches.append(match)
        
        # sort the matches by weight
        weighted_fillable = []
        for match in fillable_matches:
            weight = self._calculate_fill_weight(match, batch_idx)
            weighted_fillable.append((match, weight))
        
        weighted_fillable.sort(key=lambda x: x[1], reverse=True)
        
        # fill the matches
        for match, weight in weighted_fillable:
            if len(batch) >= self.total_court:
                break
            
            match_players = self._get_match_players(match)
            if not (match_players & batch_players):
                # insert the match into the correct position
                self._insert_match_into_batch(match, batch_idx)
                batch_players.update(match_players)
                # remove the match from the remaining matches
                remaining_matches.remove(match)

    def _can_fill_match_in_batch(self, match, batch_players):
        """check if the match can be filled into the batch"""
        # check the player conflict
        match_players = self._get_match_players(match)
        if match_players & batch_players:
            return False
        
        # check the dependency
        if not self._can_schedule_match(match):
            return False
        
        return True

    def _calculate_fill_weight(self, match, batch_idx):
        """calculate the weight of filling the match"""
        weight = 0
        
        # basic weight
        weight += self._calculate_weight(match)
        
        # additional consideration: penalty for player conflict in the batch
        batch_players = self._get_batch_players(batch_idx)
        match_players = self._get_match_players(match)
        
        if match_players & batch_players:
            weight -= 1000
        
        return weight

    def _insert_match_into_batch(self, match, batch_idx):
        """insert match into the specified batch"""
        # calculate the insert position
        insert_position = batch_idx * self.total_court + len(self._get_batch_matches(batch_idx))
        
        # insert into the correct position of scheduled_matches
        self.scheduled_matches.insert(insert_position, match)

    def _get_batch_matches(self, batch_idx):
        """get matches in the specified batch"""
        start_idx = batch_idx * self.total_court
        end_idx = start_idx + self.total_court
        return self.scheduled_matches[start_idx:end_idx]

    def _get_batch_players(self, batch_idx):
        """get players in the specified batch
        
        return a set of player.name in specific batch
        """
        batch_matches = self._get_batch_matches(batch_idx)
        players = set()
        
        for match in batch_matches:
            match_players = self._get_match_players(match)
            players.update(match_players)
        
        return players

    def _write_schedule(self, filename):
        """write schedule to Excel file, including color markers and stats"""
        try:
            rows = []
            all_consecutive_players = []
            
            # reorganize scheduled_matches into batches
            batches = self._organize_matches_into_batches()
            # print(f"Debug: batches: {batches}")
            
            # 如果沒有已安排的比賽，創建一個狀態報告
            if not batches:
                # print("Debug: No batches found, creating status report")
                rows.append({
                    'Match_ID': 'Schedule Status',
                    'Batch': 'Report',
                    'Court': 'Total Courts',
                    'Date': 'Available',
                    'Start_Time': 'Matches',
                    'End_Time': 'Scheduled',
                    'Match_Type': 'Unscheduled',
                    'Category': 'Matches',
                    'Group': 'Total',
                    'Player1/Team1': 'Matches',
                    'Player2/Team2': 'Found',
                    'Consecutive_Players': 'In',
                    'Status': 'Database',
                    'Score1': 'Please',
                    'Score2': 'Check',
                    'Umpire': 'Database',
                    'Notes': f'Total matches: {len(self.all_matches) if hasattr(self, "all_matches") else 0}, Scheduled: {len(self.scheduled_matches) if hasattr(self, "scheduled_matches") else 0}'
                })
            else:
                # print(f"Debug: Processing {len(batches)} batches")
                # 處理已安排的比賽
                for batch_num, batch in batches.items():
                    # print(f"Debug: Processing batch {batch_num} with {len(batch)} matches")
                    if batch_num == 'Unscheduled':
                        continue
                        
                    batch_idx = batch_num
                    batch_rows = []
                    
                    # add actual matches
                    for court_idx, match_info in enumerate(batch, 1):
                        # print(f"Debug: Processing match {court_idx} in batch {batch_num}")
                        # 現在 match_info 是字典 {'match': match, 'schedule_item': schedule_item}
                        match = match_info['match']
                        schedule_item = match_info['schedule_item']
                        
                        if schedule_item:
                            # print(f"Debug: Processing match {match.id} with schedule_item")
                            # get match info (using Match object attributes)
                            category = match.event_type
                            group = Group.query.filter_by(id=match.group_id).first()
                            flight = group.name if group else ''
                            
                            # get player info
                            if match.event_type in ['MD', 'WD', 'XD']:
                                # double
                                team1_p1 = match.team1_player1_name or ''
                                team1_p2 = match.team1_player2_name or ''
                                team2_p1 = match.team2_player1_name or ''
                                team2_p2 = match.team2_player2_name or ''
                                
                                # handle Winner of Match
                                if 'Winner of Match' in team1_p1:
                                    player1s = f"Winner of {team1_p1}"
                                else:
                                    player1s = f"{team1_p1} / {team1_p2}" if team1_p1 and team1_p2 else ''
                                
                                if 'Winner of Match' in team2_p1:
                                    player2s = f"Winner of {team2_p1}"
                                else:
                                    player2s = f"{team2_p1} / {team2_p2}" if team2_p1 and team2_p2 else ''
                                
                                match_type = "Double"
                            else:
                                # single atches
                                player1_name = match.player1_name or ''
                                player2_name = match.player2_name or ''
                                
                                # handle Winner of Match
                                if 'Winner of Match' in player1_name:
                                    player1s = f"Winner of {player1_name}"
                                else:
                                    player1s = player1_name
                                
                                if 'Winner of Match' in player2_name:
                                    player2s = f"Winner of {player2_name}"
                                else:
                                    player2s = player2_name
                                
                                match_type = "Single"
                            
                            # check consecutive players
                            consecutive_players = []
                            if not self._is_winner_match(match):
                                consecutive_players = self._get_consecutive_players_for_match(match, batch_idx)
                                consecutive_str = ", ".join(consecutive_players) if consecutive_players else ""
                                all_consecutive_players.extend(consecutive_players)
                            else:
                                consecutive_str = ""
                            
                            batch_rows.append({
                                'Schedule_Item_ID': schedule_item.id,
                                'Batch': batch_idx,
                                'Court': court_idx,
                                'Date': schedule_item.scheduled_date.strftime('%Y-%m-%d'),
                                'Start_Time': schedule_item.scheduled_start_time.strftime('%H:%M'),
                                'End_Time': schedule_item.scheduled_end_time.strftime('%H:%M'),
                                'Match_ID': match.id,
                                'Match_Type': match_type,
                                'Category': category,
                                'Group': flight,
                                'Player1/Team1': player1s,
                                'Player2/Team2': player2s,
                                'Consecutive_Players': consecutive_str,
                                'Status': match.status or 'Scheduled',
                                'Score1': match.player1_score or 0,
                                'Score2': match.player2_score or 0,
                                'Umpire': '',
                                'Notes': ''
                            })
                        else:
                            # print(f"Debug: Match {match.id} has no schedule_item")
                            batch_rows.append({
                                'Schedule_Item_ID': '',
                                'Batch': batch_idx,
                                'Court': court_idx,
                                'Date': '',
                                'Start_Time': '',
                                'End_Time': '',
                                'Match_ID': match.id,
                                'Match_Type': '',
                                'Category': '',
                                'Group': '',
                                'Player1/Team1': '',
                                'Player2/Team2': '',
                                'Consecutive_Players': '',
                                'Status': '',
                                'Score1': '',
                                'Score2': '',
                                'Umpire': '',
                                'Notes': ''
                            })
                    
                    # fill empty rows to reach the total_court number
                    for court_idx in range(len(batch) + 1, self.total_court + 1):
                        batch_rows.append({
                            'Schedule_Item_ID': '',
                            'Batch': batch_idx,
                            'Court': court_idx,
                            'Date': '',
                            'Start_Time': '',
                            'End_Time': '',
                            'Match_ID': '',
                            'Match_Type': '',
                            'Category': '',
                            'Group': '',
                            'Player1/Team1': '',
                            'Player2/Team2': '',
                            'Consecutive_Players': '',
                            'Status': '',
                            'Score1': '',
                            'Score2': '',
                            'Umpire': '',
                            'Notes': ''
                        })
                    
                    # add all rows of this batch to the total rows list
                    rows.extend(batch_rows)

            # 處理未安排的比賽
            if 'Unscheduled' in batches:
                unscheduled_batch = batches['Unscheduled']
                batch_rows = []
                
                for court_idx, match_info in enumerate(unscheduled_batch, 1):
                    match = match_info['match']
                    
                    # 處理未安排的比賽
                    batch_rows.append({
                        'Schedule_Item_ID': '',
                        'Batch': 'Unscheduled',
                        'Court': court_idx,
                        'Date': 'TBD',
                        'Start_Time': 'TBD',
                        'End_Time': 'TBD',
                        'Match_ID': match.id,
                        'Match_Type': 'Single' if match.event_type in ['MS', 'WS'] else 'Double',
                        'Category': match.event_type,
                        'Group': 'TBD',
                        'Player1/Team1': match.player1_name or 'TBD',
                        'Player2/Team2': match.player2_name or 'TBD',
                        'Consecutive_Players': '',
                        'Status': 'Unscheduled',
                        'Score1': 0,
                        'Score2': 0,
                        'Umpire': '',
                        'Notes': 'Not scheduled'
                    })
                
                if batch_rows:
                    rows.append({})  # empty row
                    rows.append({
                        'Schedule_Item_ID': '',
                        'Batch': '',
                        'Court': '',
                        'Date': '',
                        'Start_Time': '',
                        'End_Time': '',
                        'Match_ID': 'Unscheduled Matches',
                        'Match_Type': '',
                        'Category': '',
                        'Group': '',
                        'Player1/Team1': '',
                        'Player2/Team2': '',
                        'Consecutive_Players': '',
                        'Status': '',
                        'Score1': '',
                        'Score2': '',
                        'Umpire': '',
                        'Notes': f'Total: {len(unscheduled_batch)} matches'
                    })
                    rows.extend(batch_rows)

            # calculate the stats of affected players
            player_counts = Counter(all_consecutive_players)
            total_affected_players = len(player_counts)
            
            # add stats info row
            rows.append({})  # empty row
            rows.append({
                'Batch': 'Stats',
                'Court': '',
                'Match_Type': '',
                'Category': '',
                'Group': '',
                'Player1/Team1': '',
                'Player2/Team2': '',
                'Consecutive_Players': '',
                'Status': '',
                'Score1': '',
                'Score2': '',
                'Umpire': '',
                'Notes': ''
            })
            
            # add the total count of affected players
            rows.append({
                'Batch': 'Total Affected Players',
                'Court': '',
                'Match_Type': '',
                'Category': '',
                'Group': '',
                'Player1/Team1': '',
                'Player2/Team2': '',
                'Consecutive_Players': '',
                'Status': '',
                'Score1': '',
                'Score2': '',
                'Umpire': '',
                'Notes': total_affected_players
            })
            
            # add the count of each affected player
            for player, count in player_counts.most_common():  # sorted by count (descending order)
                rows.append({
                    'Batch': '',
                    'Court': '',
                    'Match_Type': '',
                    'Category': '',
                    'Group': '',
                    'Player1/Team1': player,
                    'Player2/Team2': '',
                    'Consecutive_Players': '',
                    'Status': '',
                    'Score1': '',
                    'Score2': '',
                    'Umpire': '',
                    'Notes': f"Consecutive {count} times"
                })

            # print(f"Debug: About to write Excel file with {len(rows)} rows")

            # Write to Excel file
            df = pd.DataFrame(rows)
            df.to_excel(filename, index=False, sheet_name='MatchSchedule')

            # print(f"Debug: Excel file written successfully")

            # Use openpyxl to add color markers
            wb = load_workbook(filename)
            ws = wb['MatchSchedule']
            
            # Yellow for marking consecutive players
            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            
            # Green for marking regular match rows
            green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            
            # Blue for marking stats info
            blue_fill = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")
            
            # Red for marking affected players
            red_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
            
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column), 2):
                consecutive_cell = row[12]  # Consecutive_Players row (修正索引)
                category_cell = row[8]      # Category row (修正索引)
                batch_cell = row[1]         # Batch row (保持不變)
                match_id_cell = row[6]      # Match_ID row (修正索引)
                
                # 調試信息
                # print(f"Row {row_idx}: Batch={batch_cell.value}, Category={category_cell.value}, Consecutive={consecutive_cell.value}, MatchID={match_id_cell.value}")
                
                # 首先檢查是否是統計行
                if batch_cell.value == 'Stats':
                    for cell in row:
                        cell.fill = blue_fill
                    # print(f"  -> Blue (Stats)")
                elif batch_cell.value == 'Total Affected Players':
                    for cell in row:
                        cell.fill = red_fill
                    # print(f"  -> Red (Total Affected Players)")
                elif batch_cell.value == '' and row[10].value and 'Consecutive' in str(row[17].value or ''):  # 修正索引
                    for cell in row:
                        cell.fill = red_fill
                    # print(f"  -> Red (Affected Players Detail)")
                # 然後檢查是否是比賽行
                elif match_id_cell.value and category_cell.value:  # 這是一個比賽行
                    if consecutive_cell.value and consecutive_cell.value.strip():  # 有 consecutive players
                        for cell in row:
                            cell.fill = yellow_fill
                        # print(f"  -> Yellow (Has Consecutive Players)")
                    else:  # 沒有 consecutive players
                        for cell in row:
                            cell.fill = green_fill
                        # print(f"  -> Green (Regular Match)")
                else:
                    # print(f"  -> No color applied")
                    pass

            wb.save(filename)
            return filename
        
        except Exception as e:
            # print(f"Error in _write_schedule: {str(e)}")
            # print(f"Error type: {type(e)}")
            import traceback
            traceback.print_exc()
            # 即使出錯，也創建一個基本的文件
            try:
                basic_data = [{
                    'Status': 'Schedule Generation Error',
                    'Message': str(e),
                    'Total Courts': self.total_court,
                    'Total Matches': len(self.all_matches) if hasattr(self, 'all_matches') else 0,
                    'Scheduled Matches': len(self.scheduled_matches) if hasattr(self, 'scheduled_matches') else 0
                }]
                df = pd.DataFrame(basic_data)
                df.to_excel(filename, index=False, sheet_name='ErrorReport')
                return filename
            except Exception as e2:
                # print(f"Error creating error report: {str(e2)}")
                raise e

    def _organize_matches_into_batches(self):
        """重新組織比賽到批次中，包含 ScheduleItem 資訊"""
        batches = {}
        
        for match in self.scheduled_matches:
            # 獲取對應的 ScheduleItem
            schedule_item = ScheduleItem.query.filter_by(match_id=match.id).first()
            
            if schedule_item:
                batch_num = schedule_item.batch_number
                if batch_num not in batches:
                    batches[batch_num] = []
                
                # 返回包含所有資訊的字典
                batches[batch_num].append({
                    'match': match,
                    'schedule_item': schedule_item
                })
        
        return batches

    def _get_consecutive_players_for_match(self, match, current_batch_idx):
        """get the consecutive players in the match"""
        consecutive_players = []
        match_players = self._get_match_players(match)
        
        # check if the player is in the previous batch
        if current_batch_idx > 1:
            previous_batch_players = self._get_players_from_batch(current_batch_idx - 1)
            for player in match_players:
                if player in previous_batch_players:
                    consecutive_players.append(player)
        
        return consecutive_players

    def _get_players_from_batch(self, batch_idx):
        """get all players in the specified batch"""
        players = set()
        batches = self._organize_matches_into_batches()
        
        # 檢查 batch_idx 是否在 batches 字典中
        if batch_idx in batches:
            batch = batches[batch_idx]
            for match_info in batch:
                match = match_info['match']
                match_players = self._get_match_players(match)
                players.update(match_players)
    
        return players
        
    def _players_not_in_selected_players(self, match, seelcted_players):
        """check if players in match are not in the selected_players set"""
        match_players = self._get_match_players(match)
        for player in match_players:
            if player in seelcted_players:
                return False
        return True

    def _create_new_batches_for_remaining(self, remaining_matches):
        """create new batches for the remaining matches, consider the dependency and player conflict"""
        if not remaining_matches:
            return
        
        # sort the remaining matches by round and match number
        remaining_matches.sort(key=lambda x: (x.round or 1, x.match_number or 1))
        
        # group the matches by round
        matches_by_round = {}
        for match in remaining_matches:
            round_num = match.round or 1
            if round_num not in matches_by_round:
                matches_by_round[round_num] = []
            matches_by_round[round_num].append(match)
        
        # process the matches by round
        for round_num in sorted(matches_by_round.keys()):
            round_matches = matches_by_round[round_num]
            self._process_round_matches(round_matches)

    def _process_round_matches(self, round_matches):
        """process the matches in a round"""
        current_batch = []
        current_batch_players = set()
        
        # compute the weight and sort the matches
        weighted_matches = []
        for match in round_matches:
            # Check if the match is scheduled
            if match not in self.scheduled_matches:
                weight = self._calculate_weight(match)
                weighted_matches.append((match, weight))
        
        weighted_matches.sort(key=lambda x: x[1], reverse=True)
        
        for match, weight in weighted_matches:
            #  check if the match is scheduled
            if match in self.scheduled_matches:
                continue
        
            # check the dependency
            if not self._can_schedule_match(match):
                continue
        
            # check player conflict
            match_players = self._get_match_players(match)
        
            if match_players & current_batch_players:
                # player conflict, start a new batch
                if current_batch:
                    self.scheduled_matches.extend(current_batch)
                    current_batch = []
                    current_batch_players = set()
        
            # check if the batch is full
            if len(current_batch) >= self.total_court:
                # batch is full, start a new batch
                self.scheduled_matches.extend(current_batch)
                current_batch = []
                current_batch_players = set()
        
            # add the match to the current batch
            current_batch.append(match)
            current_batch_players.update(match_players)
        
        # handle the last incomplete batch
        if current_batch:
            self.scheduled_matches.extend(current_batch)
    
    def _can_schedule_match(self, match):
        """check if the match can be scheduled (elimination dependency)"""
        # 跳過 BYE 比賽，不納入 schedule
        if self._is_bye_match(match):
            return False  # 改回 False，跳過 BYE 比賽
        
        if not match.round or match.round == 1:
            return True
        
        # check if the previous match exists
        prev_match1 = self._get_match_by_id(match.prev_match1_id)
        prev_match2 = self._get_match_by_id(match.prev_match2_id)
        
        # if the previous match does not exist, it might be the first round match
        if not prev_match1 or not prev_match2:
            return True
        
        # check if the previous match is scheduled
        prev_match1_scheduled = prev_match1 in self.scheduled_matches
        prev_match2_scheduled = prev_match2 in self.scheduled_matches
        
        return prev_match1_scheduled and prev_match2_scheduled

    def _is_bye_match(self, match):
        """check if the match is a bye match"""
        return (match.player1_name == 'BYE' or match.player2_name == 'BYE' or
            match.team1_player1_name == 'BYE' or match.team1_player2_name == 'BYE' or
            match.team2_player1_name == 'BYE' or match.team2_player2_name == 'BYE')

    def _is_winner_match(self, match):
        """Check ifi the match is a winner match（match info including 'Winner of Match'）"""
        if match.player1_name and 'Winner of Match' in match.player1_name:
            return True
        if match.player2_name and 'Winner of Match' in match.player2_name:
            return True
        if match.team1_player1_name and 'Winner of Match' in match.team1_player1_name:
            return True
        if match.team1_player2_name and 'Winner of Match' in match.team1_player2_name:
            return True
        if match.team2_player1_name and 'Winner of Match' in match.team2_player1_name:
            return True
        if match.team2_player2_name and 'Winner of Match' in match.team2_player2_name:
            return True
        return False

    def create_schedule(self, tournament_id, schedule_data):
        """create the schedule in the database"""
        try:
            tournament = Tournament.query.filter_by(id=tournament_id).first()
            if not tournament:
                raise ValueError(f"Tournament {tournament_id} not found")
            
            # Delete the previous match_schedule, and related objects
            existing_schedules = Schedule.query.filter_by(tournament_id=tournament_id).all()
            for existing_schedule in existing_schedules:
                # Delete related schedule items
                ScheduleItem.query.filter_by(schedule_id=existing_schedule.id).delete()
                db.session.delete(existing_schedule)
            
            db.session.commit()

            # get start_time, end_time, match_duration from schedule_data
            start_time_str = schedule_data.get('start_time', '09:00')
            end_time_str = schedule_data.get('end_time', '18:00')
            match_duration = schedule_data.get('match_duration', 30)
            
            # convert time string to time object
            start_time = datetime.strptime(start_time_str, '%H:%M').time()
            end_time = datetime.strptime(end_time_str, '%H:%M').time()
            
            # Create new schedue with fixed start_time
            schedule_dict = {
                'tournament_id': tournament_id,
                'start_date': tournament.start_date,
                'end_date': tournament.end_date,
                'start_time': start_time,
                'end_time': end_time,
                'total_courts': self.total_court,
                'match_duration': match_duration,  
                'total_matches': len(self.scheduled_matches),
                'total_batches': len(self.scheduled_matches) // self.total_court + (1 if len(self.scheduled_matches) % self.total_court > 0 else 0),
                'status': 'active'
            }
            
            # create schedule object
            schedule = Schedule(**schedule_dict)
            db.session.add(schedule)
            db.session.flush()  # flush() to get schedule.id
            
            # create schedule items
            current_date = schedule.start_date
            batch_number = 1
            order_in_batch = 0
            
            # compute time for each batch
            batch_start_time = schedule.start_time  # 9:00
            batch_end_time = datetime.combine(datetime.today(), batch_start_time) + timedelta(minutes=schedule.match_duration)
            batch_end_time = batch_end_time.time()
            
            for i, match in enumerate(self.scheduled_matches):
                # check if need a new batch
                if order_in_batch >= self.total_court:
                    batch_number += 1
                    order_in_batch = 0
                    
                    # compute the time for next batch
                    batch_start_time = batch_end_time
                    batch_end_time = datetime.combine(datetime.today(), batch_start_time) + timedelta(minutes=schedule.match_duration)
                    batch_end_time = batch_end_time.time()
                    
                    # check if the time exceed endtime, if so, need to process at the next day
                    if batch_start_time >= schedule.end_time:
                        current_date += timedelta(days=1)
                        batch_start_time = schedule.start_time
                        batch_end_time = datetime.combine(datetime.today(), batch_start_time) + timedelta(minutes=schedule.match_duration)
                        batch_end_time = batch_end_time.time()
                
                # if current date exceed tournament.end_date
                if current_date > schedule.end_date:
                    break
                
                # prepare for schedule_item_dict
                schedule_item_dict = {
                    'schedule_id': schedule.id,
                    'match_id': match.id,
                    'batch_number': batch_number,
                    'order_in_batch': order_in_batch,
                    'court_number': order_in_batch + 1,
                    'scheduled_date': current_date,
                    'scheduled_start_time': datetime.combine(current_date, batch_start_time),
                    'scheduled_end_time': datetime.combine(current_date, batch_end_time),
                    'status': 'scheduled'
                }
                
                # create ScheduleItem object
                schedule_item = ScheduleItem(**schedule_item_dict)
                db.session.add(schedule_item)
                
                order_in_batch += 1
            
            db.session.commit()

            # 暫時跳過 BYE 匹配處理
            # try:
            #     TournamentService.process_bye_matches_after_schedule(tournament_id)
            # except Exception as e:
            #     print(f"Error processing bye matches after schedule: {e}")
            #     db.session.rollback()
            #     raise e

            return schedule.id
            
        except Exception as e:
            db.session.rollback()
            raise e

    def _add_time(self, time, minutes):
        """processing addition in time"""
        dt = datetime.combine(datetime.today(), time)
        new_dt = dt + timedelta(minutes=minutes)
        return new_dt.time()

    def update_schedule_item(self, file):
        """user upload the schedule, update the schedule_items"""
        excel_data = pd.read_excel(file, engine='openpyxl')
        # for row in excel_data:
        #     schedule_item = ScheduleItem.query.filter_by(schedule_id=)

    def process_uploaded_schedule(self, file, tournament_id):
        """處理上傳的 Excel 檔案 - 包含讀取、驗證和更新"""
        try:
            # 1. 讀取 Excel 檔案
            read_result = self._read_excel_file(file)
            if not read_result['success']:
                return read_result
            
            excel_data = read_result['data']
            
            # 2. 驗證格式
            validation_result = self.validate_excel_format(excel_data)
            if not validation_result['is_valid']:
                return {
                    'status': 'error',
                    'message': 'Excel format validation failed',
                    'details': validation_result
                }
            
            # 3. 驗證 tournament 存在
            tournament = Tournament.query.get(tournament_id)
            if not tournament:
                return {
                    'status': 'error',
                    'message': 'Tournament not found'
                }
            
            # 4. 更新賽程表
            update_result = self.update_schedule_from_excel(tournament_id, excel_data)
            
            return {
                'status': update_result['status'],
                'message': 'Schedule processed successfully' if update_result['status'] == 'success' else 'Schedule processed with some issues',
                'details': update_result
            }
            
        except Exception as e:
            return {
                'status': 'error',
                'message': f'Processing failed: {str(e)}'
            }

    def _read_excel_file(self, file):
        """讀取 Excel 檔案"""
        try:
            # 檢查檔案
            if not file or file.filename == '':
                return {'success': False, 'message': 'No file selected'}
            
            if not file.filename.endswith('.xlsx'):
                return {'success': False, 'message': 'Please upload .xlsx file'}
            
            # 讀取 Excel
            df = pd.read_excel(file)
            excel_data = df.to_dict('records')
            
            if not excel_data:
                return {'success': False, 'message': 'Excel file is empty'}
            
            return {
                'success': True,
                'data': excel_data,
                'columns': list(df.columns)
            }
            
        except Exception as e:
            return {
                'success': False,
                'message': f'Error reading Excel file: {str(e)}'
            }

    def validate_excel_format(self, excel_data):
        """驗證 Excel 格式和基本資料"""
        errors = []
        warnings = []
        
        # 檢查必要欄位
        required_columns = ['Schedule_Item_ID', 'Match_ID', 'Batch', 'Court', 'Date', 'Start_Time', 'End_Time']
        
        if not excel_data:
            errors.append("Excel file is empty")
            return {'errors': errors, 'warnings': warnings, 'is_valid': False}
        
        # 檢查第一行的欄位
        first_row = excel_data[0]
        missing_columns = [col for col in required_columns if col not in first_row]
        if missing_columns:
            errors.append(f"Missing required columns: {missing_columns}")
        
        # 檢查資料完整性
        for row_idx, row in enumerate(excel_data, 1):
            if not row.get('Schedule_Item_ID'):
                warnings.append(f"Row {row_idx}: Missing Schedule_Item_ID")
            
            if not row.get('Match_ID'):
                warnings.append(f"Row {row_idx}: Missing Match_ID")
        
        return {
            'errors': errors,
            'warnings': warnings,
            'is_valid': len(errors) == 0
        }

    def update_schedule_from_excel(self, tournament_id, excel_data):
        """從 Excel 資料更新賽程表 - 處理 Match_ID 交換"""
        try:
            updated_count = 0
            errors = []
            
            for row_idx, row in enumerate(excel_data, 1):
                try:
                    schedule_item_id = row.get('Schedule_Item_ID')
                    
                    # 跳過空的 Schedule_Item_ID
                    if not schedule_item_id or str(schedule_item_id).lower() == 'nan':
                        continue
                    
                    # 查找對應的 ScheduleItem
                    schedule_item = ScheduleItem.query.get(schedule_item_id)
                    if not schedule_item:
                        errors.append(f"Row {row_idx}: ScheduleItem ID {schedule_item_id} not found")
                        continue
                    
                    # 驗證是否屬於正確的 tournament
                    if schedule_item.schedule.tournament_id != tournament_id:
                        errors.append(f"Row {row_idx}: ScheduleItem {schedule_item_id} doesn't belong to tournament {tournament_id}")
                        continue
                    
                    # 更新基本資訊（這些不應該改變，但我們可以驗證）
                    if 'Batch' in row and row['Batch'] and str(row['Batch']).strip():
                        try:
                            batch_num = int(row['Batch'])
                            if schedule_item.batch_number != batch_num:
                                errors.append(f"Row {row_idx}: Batch number mismatch. Expected {schedule_item.batch_number}, got {batch_num}")
                                continue
                        except (ValueError, TypeError):
                            errors.append(f"Row {row_idx}: Invalid batch number '{row['Batch']}'")
                            continue
                    
                    if 'Court' in row and row['Court'] and str(row['Court']).strip():
                        try:
                            court_num = int(row['Court'])
                            if schedule_item.court_number != court_num:
                                errors.append(f"Row {row_idx}: Court number mismatch. Expected {schedule_item.court_number}, got {court_num}")
                                continue
                        except (ValueError, TypeError):
                            errors.append(f"Row {row_idx}: Invalid court number '{row['Court']}'")
                            continue
                    
                    # 驗證日期和時間（修正時間比較邏輯）
                    if 'Date' in row and row['Date']:
                        try:
                            if isinstance(row['Date'], str):
                                new_date = datetime.strptime(row['Date'], '%Y-%m-%d').date()
                            else:
                                new_date = row['Date']
                        
                            if schedule_item.scheduled_date != new_date:
                                errors.append(f"Row {row_idx}: Date mismatch. Expected {schedule_item.scheduled_date}, got {new_date}")
                                continue
                        except (ValueError, TypeError) as e:
                            errors.append(f"Row {row_idx}: Invalid date format '{row['Date']}': {str(e)}")
                            continue
                    
                    if 'Start_Time' in row and row['Start_Time']:
                        try:
                            if isinstance(row['Start_Time'], str):
                                new_start_time = datetime.strptime(row['Start_Time'], '%H:%M').time()
                            else:
                                new_start_time = row['Start_Time']
                        
                            # 修正：比較時間部分，而不是完整的 datetime
                            if schedule_item.scheduled_start_time:
                                if isinstance(schedule_item.scheduled_start_time, datetime):
                                    db_start_time = schedule_item.scheduled_start_time.time()
                                else:
                                    db_start_time = schedule_item.scheduled_start_time
                                
                                if db_start_time != new_start_time:
                                    errors.append(f"Row {row_idx}: Start time mismatch. Expected {db_start_time}, got {new_start_time}")
                                    continue
                        except (ValueError, TypeError) as e:
                            errors.append(f"Row {row_idx}: Invalid start time format '{row['Start_Time']}': {str(e)}")
                            continue
                    
                    if 'End_Time' in row and row['End_Time']:
                        try:
                            if isinstance(row['End_Time'], str):
                                new_end_time = datetime.strptime(row['End_Time'], '%H:%M').time()
                            else:
                                new_end_time = row['End_Time']
                        
                            # 修正：比較時間部分，而不是完整的 datetime
                            if schedule_item.scheduled_end_time:
                                if isinstance(schedule_item.scheduled_end_time, datetime):
                                    db_end_time = schedule_item.scheduled_end_time.time()
                                else:
                                    db_end_time = schedule_item.scheduled_end_time
                                
                                if db_end_time != new_end_time:
                                    errors.append(f"Row {row_idx}: End time mismatch. Expected {db_end_time}, got {new_end_time}")
                                    continue
                        except (ValueError, TypeError) as e:
                            errors.append(f"Row {row_idx}: Invalid end time format '{row['End_Time']}': {str(e)}")
                            continue
                    
                    # 更新 Match_ID（這是主要的變更）
                    if 'Match_ID' in row and row['Match_ID']:
                        try:
                            new_match_id = int(row['Match_ID'])
                            if schedule_item.match_id != new_match_id:
                                # 驗證新的 Match_ID 是否存在
                                new_match = Match.query.get(new_match_id)
                                if not new_match:
                                    errors.append(f"Row {row_idx}: Match ID {new_match_id} not found")
                                    continue
                                
                                # 驗證新的 Match 是否屬於同一個 tournament
                                if new_match.tournament_id != tournament_id:
                                    errors.append(f"Row {row_idx}: Match ID {new_match_id} doesn't belong to tournament {tournament_id}")
                                    continue
                                
                                # 更新 Match_ID
                                schedule_item.match_id = new_match_id
                        except (ValueError, TypeError):
                            errors.append(f"Row {row_idx}: Invalid Match_ID '{row['Match_ID']}'")
                            continue
                    
                    # 更新 updated_at 時間戳
                    schedule_item.updated_at = datetime.utcnow()
                    
                    updated_count += 1
                    
                except Exception as e:
                    errors.append(f"Row {row_idx}: Unexpected error: {str(e)}")
                    continue
            
            # 提交所有更改
            if updated_count > 0:
                db.session.commit()
            
            return {
                'status': 'success' if not errors else 'partial_success',
                'updated_items': updated_count,
                'total_items': len(excel_data),
                'errors': errors
            }
            
        except Exception as e:
            db.session.rollback()
            return {
                'status': 'error',
                'message': f'Database error: {str(e)}',
                'updated_items': 0,
                'total_items': len(excel_data)
            }